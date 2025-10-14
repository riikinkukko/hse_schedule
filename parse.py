import requests
import json
from re import findall
from time import sleep
from os.path import join
from os import makedirs
from openpyxl.reader.excel import load_workbook

from pprint import pprint


MAIN_ID = '1BMR4Zk3BU2Tyo7L-CYJeMfVuCxjmmT94kfz4Jp6BFAc'
MAIN_GID = 739453176
ENGLISH_ID = '1RB9AWtrYm6Y9m8NSy6On7Zk3byws8RonAGBqeneSxOo'
ENGLISH_GID = 23993546
GROUP_NUMBER = 5
DEST_FOLDER = ''
OUTPUT_FILE = 'schelude.json'


def check_for_redirect(response):
    if response.history:
        raise requests.HTTPError


def get_response(url):
    while True:
        try:
            response = requests.get(url)
            response.raise_for_status()
            check_for_redirect(response)
            return response
        except ConnectionError:
            print('Подключение отсутствует')
            sleep(15)


def download_timetable(id, gid, timetable_filename, dest_folder=''):
    url = f'https://docs.google.com/spreadsheets/d/{id}/export'
    response = requests.get(url, params={'gid': gid})
    if dest_folder:
        makedirs(dest_folder, exist_ok=True)
    filepath = join(dest_folder, timetable_filename).replace(r'\\', '/')
    with open(filepath, 'wb') as file:
        file.write(response.content)
    return filepath


def format_lessons(value, classnumber):
    value = value.replace('\n', '').replace('-'*21, '').replace('-'*9, '').strip()
    value = value.rsplit(' - ', maxsplit=1)
    if len(value) > 1 and classnumber:
        lecture, teacher = value
        lesson = {
            'lesson_name': lecture,
            'teacher': teacher,
            'classnumber': int(classnumber) if classnumber != '-' else 'online'
        }
        return lesson
    return 'None'


def save_json(data, filename, folder=''):
    if folder:
        makedirs(folder, exist_ok=True)
    filepath = join(folder, filename)
    with open(filepath, 'w', encoding='utf8') as file:
        json.dump(data, file, ensure_ascii=False)
    return filepath


def get_data_from_main_xlsx(filepath):
    groups = ['E', 'I', 'M', 'Q', 'U', 'Y', 'AC']
    schelude = {
        'mon': [],
        'tue': [],
        'wed': [],
        'thu': [],
        'fri': [],
        'sat': [],
    }
    sheet = load_workbook(filename=filepath, data_only=True).active

    line = 11
    for key in schelude.keys():
        for _ in range(9):
            line += 1
            value = sheet[f'{groups[GROUP_NUMBER-1]}{line}'].value
            classnumber = sheet[f'{chr(ord(groups[GROUP_NUMBER-1]) + 1)}{line}'].value
            schelude[key].append(format_lessons(value, classnumber) if value else 'None')
    return schelude


def del_spaces(array):
    return [x for x in array if x != '']


def format_eng_lessons(groups, teachers, classnumbers):
    lessons = []
    for i in range(len(groups)):
        lesson = {
            'lesson_name': 'Английский язык',
            'teacher': teachers[i],
            'classnumber': classnumbers[i],
            'group': [int(i) for i in findall(r'\d+', groups[i])][0],
        }
        lessons.append(lesson)
    return lessons


def get_data_from_eng_xlsx(filepath):
    faculties = {
        'gym': 'F',
        'knt': 'L',
        'mng': 'R',
        'prv': 'X',
    }
    schelude = {
        'mon': [],
        'tue': [],
        'wed': [],
        'thu': [],
        'fri': [],
        'sat': [],
    }
    sheet = load_workbook(filename=filepath, data_only=True).active
    faculty = faculties['knt']
    line = 10
    for key in schelude.keys():
        for _ in range(7):
            line += 1
            column_groups = sheet[f'{faculty}{line}'].value
            if not column_groups:
                schelude[key].append('None')
                continue
            column_teachers = sheet[f'{chr(ord(faculty) + 1)}{line}'].value
            column_classnumbers = sheet[f'{chr(ord(faculty) + 2)}{line}'].value
            groups = del_spaces(list(map(str.strip, column_groups.split('\n'))))
            teachers = del_spaces(list(map(str.strip, column_teachers.split('\n'))))
            classnumbers = del_spaces(list(map(str.strip, column_classnumbers.split('\n'))))
            lessons = format_eng_lessons(groups, teachers, classnumbers)
            schelude[key].append(lessons)
    return schelude


def join_scheludes(eng_schelude, schelude):
    for key in schelude.keys():
        if eng_schelude[key] != ['None']*7:
            schelude[key] = eng_schelude[key]
    return schelude


def main():
    schelude = get_data_from_main_xlsx(download_timetable(MAIN_ID, MAIN_GID, 'timetable.xlsx'))
    eng_schelude = get_data_from_eng_xlsx(download_timetable(ENGLISH_ID, ENGLISH_GID, 'timetable_eng.xlsx'))
    data = join_scheludes(eng_schelude, schelude)
    pprint(data, sort_dicts=False)
    save_json(data, OUTPUT_FILE)


if __name__ == '__main__':
    main()
