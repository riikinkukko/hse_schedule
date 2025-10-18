import aiohttp
import asyncio
import json
from re import findall
import os
from os.path import join
from os import makedirs
from openpyxl.reader.excel import load_workbook

from database.database import db

MAIN_ID = '1BMR4Zk3BU2Tyo7L-CYJeMfVuCxjmmT94kfz4Jp6BFAc'
MAIN_GID = 739453176
ENGLISH_ID = '1RB9AWtrYm6Y9m8NSy6On7Zk3byws8RonAGBqeneSxOo'
ENGLISH_GID = 23993546
DEST_FOLDER = ''
OUTPUT_FILE = 'schelude.json'


async def check_for_redirect(response):
    if response.history:
        raise aiohttp.ClientError("Request was redirected")


async def get_response(session, url):
    while True:
        try:
            async with session.get(url) as response:
                response.raise_for_status()
                await check_for_redirect(response)
                return response
        except (aiohttp.ClientError, asyncio.TimeoutError):
            print('Подключение отсутствует, повторная попытка через 15 секунд...')
            await asyncio.sleep(15)


async def download_timetable(session, id, gid, timetable_filename, dest_folder=''):
    url = f'https://docs.google.com/spreadsheets/d/{id}/export'
    async with session.get(url, params={'gid': gid}) as response:
        if dest_folder:
            makedirs(dest_folder, exist_ok=True)
        filepath = join(dest_folder, timetable_filename).replace(r'\\', '/')
        content = await response.read()
        with open(filepath, 'wb') as file:
            file.write(content)
        return filepath


def format_lessons(value, classnumber, group_cst):
    if not value:
        return 'None'

    if isinstance(value, (int, float)):
        value = str(value)

    value = value.replace('\n', '').replace('-' * 21, '').replace('-' * 9, '').strip()

    if classnumber:
        if isinstance(classnumber, (int, float)):
            classnumber = str(classnumber)

        classnumber_parts = classnumber.split('\n')
        if classnumber_parts:
            classnumber = classnumber_parts[0].strip()
        classnumber = classnumber.replace('---', '').strip()

    value = value.rsplit(' - ', maxsplit=1)
    if len(value) > 1 and classnumber:
        lecture, teacher = value
        try:
            if isinstance(classnumber, str) and (classnumber.lower() == 'online' or classnumber == '-'):
                classroom = 'online'
            else:
                # Преобразуем в строку для обработки
                classnumber_str = str(classnumber) if not isinstance(classnumber, str) else classnumber
                # Берем только цифры из номера аудитории
                numbers = ''.join(filter(str.isdigit, classnumber_str))
                classroom = int(numbers) if numbers else 'online'
        except (ValueError, TypeError):
            classroom = 'online'

        lesson = {
            'lesson_name': lecture,
            'teacher': teacher,
            'classnumber': classroom
        }
        return lesson
    return 'None'


def save_json(data, filename, folder=''):
    if folder:
        makedirs(folder, exist_ok=True)
    filepath = join(folder, filename)
    with open(filepath, 'w', encoding='utf8') as file:
        json.dump(data, file, ensure_ascii=False)
    return filepath


def get_data_from_main_xlsx(filepath, group_cst):
    groups = ['E', 'I', 'M', 'Q', 'U', 'Y', 'AC']
    schelude = {
        'mon': [], 'tue': [], 'wed': [], 'thu': [], 'fri': [], 'sat': [],
    }
    sheet = load_workbook(filename=filepath, data_only=True).active

    line = 11
    for key in schelude.keys():
        for _ in range(9):
            line += 1
            value = sheet[f'{groups[group_cst - 1]}{line}'].value
            classnumber = sheet[f'{chr(ord(groups[group_cst - 1]) + 1)}{line}'].value
            schelude[key].append(format_lessons(value, classnumber, group_cst) if value else 'None')
    return schelude


def del_spaces(array):
    return [x for x in array if x != '']


def format_eng_lessons(groups, teachers, classnumbers, user_eng_group):
    lessons = []
    for i in range(len(groups)):
        group_numbers = [int(num) for num in findall(r'\d+', groups[i])]
        if user_eng_group in group_numbers:
            classroom = classnumbers[i]
            if classroom:
                if isinstance(classroom, (int, float)):
                    classroom = str(classroom)

                if isinstance(classroom, str):
                    classroom_parts = classroom.split('\n')
                    if classroom_parts:
                        classroom = classroom_parts[0].strip()
                    classroom = classroom.replace('---', '').strip()

                    try:
                        if classroom.lower() == 'online' or classroom == '-':
                            classroom = 'online'
                        else:
                            numbers = ''.join(filter(str.isdigit, classroom))
                            classroom = int(numbers) if numbers else classroom
                    except (ValueError, TypeError):
                        classroom = 'online'
                else:
                    classroom = str(classroom) if classroom else 'online'
            else:
                classroom = 'online'

            lesson = {
                'lesson_name': 'Английский язык',
                'teacher': teachers[i],
                'classnumber': classroom,
                'group': user_eng_group,
            }
            lessons.append(lesson)
    return lessons if lessons else 'None'


def get_data_from_eng_xlsx(filepath, user_eng_group):
    faculties = {'gym': 'F', 'knt': 'L', 'mng': 'R', 'prv': 'X'}
    schelude = {
        'mon': [], 'tue': [], 'wed': [], 'thu': [], 'fri': [], 'sat': [],
    }

    try:
        workbook = load_workbook(filename=filepath, data_only=True)
        sheet = workbook.active
        faculty = faculties['knt']
        line = 10

        for key in schelude.keys():
            for _ in range(7):
                line += 1
                column_groups = sheet[f'{faculty}{line}'].value
                if not column_groups:
                    schelude[key].append('None')
                    continue

                column_teachers = sheet[f'{chr(ord(faculty) + 1)}{line}'].value
                column_classnumbers = sheet[f'{chr(ord(faculty) + 2)}{line}'].value

                groups_str = str(column_groups) if column_groups is not None else ""
                teachers_str = str(column_teachers) if column_teachers is not None else ""
                classnumbers_str = str(column_classnumbers) if column_classnumbers is not None else ""

                groups = del_spaces(list(map(str.strip, groups_str.split('\n'))))
                teachers = del_spaces(list(map(str.strip, teachers_str.split('\n'))))
                classnumbers = del_spaces(list(map(str.strip, classnumbers_str.split('\n'))))

                min_length = min(len(groups), len(teachers), len(classnumbers))
                if min_length > 0:
                    groups = groups[:min_length]
                    teachers = teachers[:min_length]
                    classnumbers = classnumbers[:min_length]
                    lessons = format_eng_lessons(groups, teachers, classnumbers, user_eng_group)
                    schelude[key].append(lessons)
                else:
                    schelude[key].append('None')

        workbook.close()
        return schelude
    except Exception as e:
        print(f"Ошибка при чтении файла английского: {e}")
        return schelude


def join_scheludes(eng_schelude, schelude):
    for key in schelude.keys():
        if eng_schelude[key] != ['None'] * 7:
            schelude[key] = eng_schelude[key]
    return schelude


async def get_schedule_data(user_id: int):
    try:
        user_groups = db.get_user_groups(user_id)
        if not user_groups:
            raise ValueError("User groups not found")

        group_cst, group_eng = user_groups
        print(f"Получение расписания для пользователя {user_id}: КНТ-{group_cst}, Англ-{group_eng}")

        async with aiohttp.ClientSession() as session:
            main_file, eng_file = await asyncio.gather(
                download_timetable(session, MAIN_ID, MAIN_GID, 'timetable.xlsx'),
                download_timetable(session, ENGLISH_ID, ENGLISH_GID, 'timetable_eng.xlsx')
            )

            print("Файлы скачаны, начинаю парсинг...")

            schelude = get_data_from_main_xlsx(main_file, group_cst)
            print("Основное расписание распарсено")

            eng_schelude = get_data_from_eng_xlsx(eng_file, group_eng)
            print("Расписание английского распарсено")

            data = join_scheludes(eng_schelude, schelude)
            print("Расписания объединены")

            try:
                if os.path.exists(main_file):
                    os.remove(main_file)
                if os.path.exists(eng_file):
                    os.remove(eng_file)
                print("Временные файлы очищены")
            except Exception as e:
                print(f"Ошибка при очистке временных файлов: {e}")

            return data
    except Exception as e:
        print(f"Ошибка в get_schedule_data: {e}")
        import traceback
        print(f"Подробности ошибки: {traceback.format_exc()}")
        raise

async def fetch_schedule(user_id: int):
    try:
        print(f"Запрос расписания для пользователя {user_id}")
        schedule_data = await get_schedule_data(user_id)
        print("Расписание успешно получено")
        return schedule_data
    except Exception as e:
        print(f"Критическая ошибка при получении расписания: {e}")
        import traceback
        print(f"Полный traceback: {traceback.format_exc()}")
        return None